import numpy as np
from qqdm import qqdm
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.transforms as transforms
import matplotlib.image as mpimg

import os, time, sys, csv, random, math
from collections import deque, namedtuple
from multiprocessing import Pool, cpu_count
import itertools

import torch
import torch.optim as optim
import torch.nn as nn
import torch.nn.functional as F
from torch.linalg import solve as solve_matrix_system


######## SET PARALLEL COMPUTING ##########
num_cores = cpu_count()
torch.set_num_interop_threads(num_cores) # Inter-op parallelism
torch.set_num_threads(num_cores) # Intra-op parallelism
##########################################


######## SET DEVICE ######################
# device = "cuda:0" if torch.cuda.is_available() else "cpu"
device = "cpu"
#########################################


######### SET SEED ######################
seed = 1
np.random.seed(seed)
random.seed(seed)
torch.manual_seed(seed)
if torch.cuda.is_available():
    torch.cuda.manual_seed(seed)
torch.backends.cudnn.deterministic = True
torch.backends.cudnn.benchmark = False
os.environ["PYTHONHASHSEED"] = str(seed)
#########################################


################# SPEED UP ####################
torch.autograd.set_detect_anomaly(False);
torch.autograd.profiler.emit_nvtx(False);
torch.autograd.profiler.profile(False);
################################################


VectorScore = namedtuple('VectorScore', ('collision', 'distance', 'oob'))

def array(*args, **kwargs):
    kwargs.setdefault("dtype", np.float32)
    return np.array(*args, **kwargs)


class Car:
    def __init__(self, position):
        self.lf = 0.5
        self.lr = 0.5
        self.max_speed = 20.0
        self.reset(position)

    def move(self, df, a):
        self.v += a
        self.v = np.maximum(-self.max_speed, np.minimum(self.max_speed, self.v))
        self.beta += np.arctan((1 + self.lr / self.lf) * np.tan(df))
        arg = self.phi + self.beta
        self.prev_position = np.copy(self.position)
        self.position += self.v * array([np.cos(arg), np.sin(arg)])
        ang = array([np.cos(self.beta), np.sin(self.beta)])
        self.prev_front = self.front
        self.front = self.position + self.lf * ang
        self.back = self.position - self.lr * ang

    def reset(self, position):
        self.position = position
        self.prev_position = np.copy(position)
        self.v = 0
        self.phi = 0
        self.beta = 0
        self.front = array([self.position[0] + self.lf, self.position[1]])
        self.back = array([self.position[0], self.position[1] + self.lr])
        self.prev_front = np.copy(self.front)


class Jaywalker:
    def __init__(self):
        self.min_j_speed = 0.1
        self.max_j_speed = 0.5
        self.jaywalker_speed = 0.0
        self.jaywalker_dir = 1
        self.env_type = 1
        self.curriculum_stage = 1
        self.completed_mean = 0.0

        try:
            self.car_img = mpimg.imread(".car/carontop.png")
        except FileNotFoundError:
            print("Warning: Car image not found. Rendering will use a placeholder.")
            self.car_img = None # Placeholder if image is not found

        self.reward_size = 3
        self.dim_x = 100
        self.dim_y = 10
        self.n_lanes = 2
        self.lane_width = self.dim_y / self.n_lanes
        self.lanes_y = [self.lane_width/2 + i * self.lane_width for i in range(self.n_lanes)]
        self.obstacles = []
        self.max_obstacles = 1
        self.goal = array([self.dim_x, self.dim_y/4])
        self.jaywalker = array([self.dim_x/2, self.dim_y/4])
        self.jaywalker_r = 2
        self.jaywalker_max = self.jaywalker + self.jaywalker_r
        self.jaywalker_min = self.jaywalker - self.jaywalker_r
        self.df = [-np.pi/18, 0, np.pi/18]
        self.a = [-7, 0, 2]
        self.actions = [p for p in itertools.product(self.df, self.a)]
        self.num_df_actions = len(self.df)
        self.num_a_actions = len(self.a)
        self.state_size = 8
        self.action_size = self.num_df_actions * self.num_a_actions
        self.max_iterations = 15000
        self.car = Car(array([0.0,2.5]))
        self.counter_iterations = 0
        self.prev_target_distance = np.linalg.norm(self.car.front - self.goal)
        self.noise = 1e-5
        self.sight = 60
        self.sight_obstacle = 80
        self.scale_factor = 100

    def collision_with_jaywalker(self):
        front = np.maximum(self.car.front, self.car.prev_front)
        prev = np.minimum(self.car.front, self.car.prev_front)
        denom = front - prev + self.noise
        upper = (self.jaywalker_max - prev) / denom
        lower = (self.jaywalker_min - prev) / denom
        scalar_upper = np.min(upper)
        scalar_lower = np.max(lower)
        return scalar_upper >= 0 and scalar_lower <= 1 and scalar_lower <= scalar_upper

    def collision_with_obstacle(self):
        car_r = 2.0
        for obs in self.obstacles:
            obs_max = obs['pos'] + obs['r']
            obs_min = obs['pos'] - obs['r']
            
            # Simple bounding box check for performance
            if (self.car.front[0] + car_r > obs_min[0] and self.car.front[0] - car_r < obs_max[0] and
                self.car.front[1] + car_r > obs_min[1] and self.car.front[1] - car_r < obs_max[1]):
                return True
        return False

    def vision(self):
        vector_to_jaywalker = self.jaywalker - self.car.position + self.noise
        distance = np.linalg.norm(vector_to_jaywalker)
        if self.car.position[0] >= self.jaywalker[0] or distance > self.sight:
            return 0, -np.pi
        angle = np.arctan2(vector_to_jaywalker[1], vector_to_jaywalker[0])
        return 1 / distance, angle

    def vision_obstacle(self):
        if not self.obstacles:
            return 0, -np.pi
        dists = [np.linalg.norm(obs['pos'] - self.car.position) for obs in self.obstacles]
        i_min = np.argmin(dists)
        obs = self.obstacles[i_min]
        vector_to_obs = obs['pos'] - self.car.position + self.noise
        distance = np.linalg.norm(vector_to_obs)
        if distance > self.sight_obstacle or obs['pos'][0] < self.car.position[0]:
            return 0, -np.pi
        angle = np.arctan2(vector_to_obs[1], vector_to_obs[0])
        return 1 / distance, angle
    
    def is_own_lane_blocked(self, threshold=15.0):
        # Simplified check
        return False

    def step(self, action):
        self.jaywalker[1] += self.jaywalker_speed * self.jaywalker_dir
        if self.jaywalker[1] < 0 or self.jaywalker[1] > self.dim_y:
            self.jaywalker[0] = random.uniform(self.dim_x/2, self.dim_x)
            self.jaywalker_dir *= -1
            self.jaywalker[1] = np.clip(self.jaywalker[1], 0, self.dim_y)
        self.jaywalker_max = self.jaywalker + self.jaywalker_r
        self.jaywalker_min = self.jaywalker - self.jaywalker_r

        df, a = self.actions[action]
        self.car.move(df, a)

        for obs in self.obstacles:
            obs['pos'][0] -= obs['v']

        reward = np.zeros(self.reward_size)
        terminated = False
        completed = False

        if self.collision_with_jaywalker() or self.collision_with_obstacle():
            reward[0] = -10
            terminated = True
        
        reward[1] = (self.car.position[0] - self.car.prev_position[0]) / self.scale_factor

        if self.car.front[0] >= self.goal[0]:
            if not terminated:
                completed = True
            terminated = True

        if not (0 <= self.car.front[1] <= self.dim_y and 0 <= self.car.back[1] <= self.dim_y and self.car.position[0] >= 0):
            reward[2] -= 1000
            terminated = True
        else:
            reward[2] = -np.abs(self.car.position[1] - self.goal[1])
        reward[2] /= self.scale_factor * 10

        inv_distance, angle = self.vision()
        inv_distance_obs, angle_obs = self.vision_obstacle()
        
        lane_idx = min(range(len(self.lanes_y)), key=lambda i: abs(self.car.position[1]-self.lanes_y[i]))
        state = array([self.car.position[1], inv_distance, angle, self.car.v, self.car.beta, inv_distance_obs, angle_obs, float(lane_idx)])
        
        self.counter_iterations += 1
        truncated = self.counter_iterations >= self.max_iterations
        
        return state, reward, terminated, truncated, completed

    def reset(self):
        self.obstacles = []
        self.car.reset(array([0.0, 2.5]))
        self.counter_iterations = 0
        self.jaywalker = array([self.dim_x * 0.5, self.dim_y / 4])
        self.jaywalker_max = self.jaywalker + self.jaywalker_r
        self.jaywalker_min = self.jaywalker - self.jaywalker_r

        pos_x, speed, self.jaywalker_speed, self.jaywalker_dir = 0, 0, 0, 0
        
        if self.env_type == 1: # Scenario 1: ostacolo distante, pedone fermo
            pos_x, speed, self.jaywalker_speed, self.jaywalker_dir = self.dim_x * 0.75, 0.5, 0.0, 0
        elif self.env_type == 2: # Scenario 2: ostacolo vicino, pedone fermo
            pos_x, speed, self.jaywalker_speed, self.jaywalker_dir = self.jaywalker[0] + 15, 1.0, 0.0, 0
        elif self.env_type == 3: # Scenario 3: ostacolo distante, pedone mobile
            pos_x, speed, self.jaywalker_speed, self.jaywalker_dir = self.dim_x * 0.75, 0.5, 1.0, random.choice([-1, 1])
        elif self.env_type == 4: # Scenario 4: ostacolo vicino, pedone mobile
            pos_x, speed, self.jaywalker_speed, self.jaywalker_dir = self.jaywalker[0] + 15, 1.0, 1.0, random.choice([-1, 1])
        
        self.obstacles.append({'type': 'car', 'pos': array([pos_x, self.lanes_y[1]]), 'r': 2.0, 'v': speed})
        
        inv_distance, angle = self.vision()
        inv_distance_obs, angle_obs = self.vision_obstacle()
        lane_idx = min(range(len(self.lanes_y)), key=lambda i: abs(self.car.position[1]-self.lanes_y[i]))
        
        return array([self.car.position[1], inv_distance, angle, self.car.v, self.car.beta, inv_distance_obs, angle_obs, float(lane_idx)])

    def random_action(self):
        return int(np.floor(random.random() * self.action_size))

    def __str__(self):
        return "jaywalker"

    def render(self):
        plt.clf()
        ax = plt.gca()
        ax.add_patch(mpatches.Rectangle((0, 0), self.dim_x, self.dim_y, facecolor='black'))
        ax.plot([self.goal[0], self.goal[0]], [0, self.dim_y], color='lime', linewidth=2, linestyle='--', label='Finish')
        plt.plot([0, self.dim_x], [0, 0], color='white', linewidth=2)
        plt.plot([0, self.dim_x], [self.dim_y, self.dim_y], color='white', linewidth=2)
        plt.plot([0, self.dim_x], [self.dim_y / 2, self.dim_y / 2], color='white', linewidth=1, linestyle='--')
        
        ax.add_patch(mpatches.Rectangle(self.jaywalker - self.jaywalker_r, 2 * self.jaywalker_r, 2 * self.jaywalker_r, color='red', alpha=0.7))
        for obs in self.obstacles:
            ax.add_patch(mpatches.Rectangle(obs['pos'] - obs['r'], 2 * obs['r'], 2 * obs['r'], color='orange', alpha=0.7))

        if self.car_img:
            car_length, car_width = 3.0, 3.0
            extent = [self.car.position[0] - car_length / 2, self.car.position[0] + car_length / 2, self.car.position[1] - car_width / 2, self.car.position[1] + car_width / 2]
            img_transform = transforms.Affine2D().rotate_around(self.car.position[0], self.car.position[1], self.car.phi + self.car.beta) + ax.transData
            plt.imshow(self.car_img, extent=extent, transform=img_transform, zorder=5)
        else: # Fallback to a simple rectangle if image is missing
            ax.add_patch(mpatches.Rectangle(self.car.position - 1, 2, 2, color='blue', zorder=5))

        plt.xlim(-1, self.dim_x+1)
        plt.ylim(-1, self.dim_y+1)
        plt.pause(0.001)

# --- Tutte le classi Network (Q_Network, Lex_Q_Network, etc.) rimangono INVARIATE ---
# ... (omesso per brevità, il codice è identico a quello fornito)

class Q_Network(nn.Module):

    def __init__(self, n_observations, hidden = 128, weights = None):
        
        super().__init__()
        self.layer1 = nn.Linear(n_observations, hidden)
        self.layer2 = nn.Linear(hidden, hidden)
        self.layer3 = nn.Linear(hidden, hidden)
        self.layer4 = nn.Linear(hidden, hidden)

        self.input_norm = nn.BatchNorm1d(n_observations)
        self.norm = nn.BatchNorm1d(hidden)

        self.criterion = nn.SmoothL1Loss()
        self.activation = F.relu

        self.weights = weights


    def instantiate_optimizer(self, learning_rate = 1e-4):
        self.optimizer = optim.SGD(self.parameters(), lr = learning_rate, momentum = 0.9, nesterov = True)

    
    def common_forward(self, x):

        x = self.input_norm(x)
        
        z = self.activation(self.layer1(x))
        
        y = self.activation(self.layer2(z))

        y = self.activation(self.layer3(y))

        z = z + self.activation(self.layer4(y))
        
        z = self.norm(z)

        return z
    
class Lex_Q_Network(Q_Network):

    def __init__(self, n_observations, n_actions, hidden = 128, learning_rate = 1e-4, weights = None):
        
        super().__init__(n_observations, hidden, weights)

        self.layerb1 = nn.Linear(hidden, n_actions)
        self.layerb2 = nn.Linear(hidden, n_actions)
        self.layerb3 = nn.Linear(hidden, n_actions)

        self.I = torch.eye(hidden, device = device)
        self.I_o1 = torch.eye(self.layerb1.out_features, device = device)
        self.I_o2 = torch.eye(self.layerb1.out_features + self.layerb2.out_features, device = device)

        self.instantiate_optimizer(learning_rate)


    def forward(self, x):

        z = self.common_forward(x)

        o1 = F.sigmoid(self.layerb1(z)) - 1
        o2 = self.layerb2(z)
        o3 = self.layerb3(z)

        return torch.stack((o1, o2, o3), dim = 2)
    
    def learn(self, predict, target):
        self.optimizer.zero_grad()
        loss_f = self.criterion(predict[:,0], target[:,0])
        loss_i1 = self.criterion(predict[:,1], target[:,1])
        loss_i2 = self.criterion(predict[:, 2], target[:,2])
        
        loss = loss_f + loss_i1 + loss_i2
        loss.backward()

        self.optimizer.step()
        
    def  __str__(self):
        return "Lex"

class QAgent():
    def __init__(self, network, env, learning_rate=1e-2, batch_size=256, hidden=128, slack=0.1, 
                 epsilon_start=1.0, epsilon_decay=0.997, epsilon_min=0.01, episodes=3000, gamma=0.95, 
                 train_start=1000, replay_frequency=3, target_model_update_rate=1e-3, 
                 memory_length=10000, mini_batches=4, weights=None, env_type=1):

        self.env = env
        self.env.env_type = env_type

        self.batch_size = batch_size
        self.state_size = env.state_size
        self.action_size = env.action_size
        self.reward_size = env.reward_size
        self.slack = slack
        self.gamma = gamma
        self.epsilon = epsilon_start # Necessario per act() durante il test
        
        self.permissible_actions = torch.tensor(range(self.action_size)).to(device)

        self.model = network(self.state_size, self.action_size, hidden, learning_rate, weights)
        self.model.to(device)
        self.model.eval() # Imposta il modello in modalità valutazione di default

    def greedy_arglexmax(self, Q):
        permissible_actions = self.permissible_actions.clone()
        
        # 1. Filtra per collisione (priorità massima)
        mask_collision = (Q[:, 0] >= -0.7)
        if torch.any(mask_collision):
            permissible_actions = permissible_actions[mask_collision]
        else:
            # Se nessuna azione sicura è prevista, scegli quella "meno peggio"
            best_collision_val = Q[:, 0].max()
            permissible_actions = permissible_actions[Q[:, 0] >= best_collision_val - self.slack]

        # 2. Filtra per progresso (seconda priorità)
        q_forward = Q[permissible_actions, 1]
        best_forward_val = q_forward.max()
        permissible_actions = permissible_actions[q_forward >= best_forward_val - self.slack * torch.abs(best_forward_val)]
        
        # 3. Filtra per stabilità in corsia (terza priorità) e scegli l'azione migliore
        q_lane = Q[permissible_actions, 2]
        best_action_idx = q_lane.argmax()
        
        return permissible_actions[best_action_idx]

    def act(self, state):
        # Durante il test, agiamo sempre in modo greedy (epsilon = 0)
        with torch.no_grad():
            Q = self.model(state).squeeze(0) # squeeze(0) per rimuovere la dimensione del batch
            return self.greedy_arglexmax(Q)

    # ===================================================================
    # NUOVA FUNZIONE DI TEST AUTOMATIZZATO
    # ===================================================================
    def test_model(self, model_path, num_episodes=100, render=True):
        """
        Testa un modello pre-addestrato per un dato numero di episodi.
        
        Args:
            model_path (str): Il percorso al file del modello (.pt).
            num_episodes (int): Il numero di episodi di test.
            render (bool): Se True, visualizza l'ambiente durante il test.

        Returns:
            float: La percentuale di successo (episodi completati con successo).
        """
        try:
            self.model.load_state_dict(torch.load(model_path, map_location=device))
            self.model.eval()
        except FileNotFoundError:
            print(f"Errore: File del modello non trovato in {model_path}")
            return 0.0
        except Exception as e:
            print(f"Errore durante il caricamento del modello {model_path}: {e}")
            return 0.0

        successful_episodes = 0
        
        # Usa qqdm per una barra di progresso pulita
        pbar = qqdm(range(num_episodes), desc=f"Testing {os.path.basename(model_path)} on Scenario {self.env.env_type}")

        for episode in pbar:
            state = self.env.reset()
            state = torch.tensor(state, dtype=torch.float32).to(device)
            done = False
            completed = False

            while not done:
                if render:
                    self.env.render()

                action = self.act(state.unsqueeze(0))
                next_state, _, terminated, truncated, completed_episode = self.env.step(action)
                
                done = terminated or truncated
                state = torch.tensor(next_state, dtype=torch.float32).to(device)
                
                if done and completed_episode:
                    completed = True

            if completed:
                successful_episodes += 1
            
            # Aggiorna la barra di progresso con le info correnti
            pbar.set_infos({
                'Success Rate': f"{100 * successful_episodes / (episode + 1):.2f}%"
            })

        success_rate_percent = (successful_episodes / num_episodes) * 100
        return success_rate_percent


# ===================================================================
# BLOCCO DI ESECUZIONE PRINCIPALE PER IL TESTING AUTOMATIZZATO
# ===================================================================
if __name__ == "__main__":
    # Import necessario per scrivere file Excel
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("La libreria 'openpyxl' è necessaria per salvare in formato Excel.")
        print("Per favore, installala eseguendo: pip install openpyxl")
        sys.exit(1)

    # --- Parametri di configurazione ---
    MODELS_DIR = "models/"
    RESULTS_EXCEL = "test_results.xlsx" # Nome del file Excel di output
    NUM_TEST_EPISODES = 50 # Numero di episodi per testare ogni scenario

    # Verifica che la cartella dei modelli esista
    if not os.path.isdir(MODELS_DIR):
        print(f"Errore: La cartella '{MODELS_DIR}' non è stata trovata.")
        print("Per favore, crea la cartella e inserisci i file dei modelli (.pt) al suo interno.")
        sys.exit(1)

    model_files = [f for f in os.listdir(MODELS_DIR) if f.endswith(".pt")]
    if not model_files:
        print(f"Nessun file modello (.pt) trovato nella cartella '{MODELS_DIR}'.")
        sys.exit(0)

    print(f"Trovati {len(model_files)} modelli da testare.")

    # Crea un nuovo Workbook Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Risultati Test Modelli"

    # Scrivi l'intestazione
    header = ['Nome Modello', 'Successo Scenario 1 (%)', 'Successo Scenario 2 (%)', 'Successo Scenario 3 (%)', 'Successo Scenario 4 (%)']
    sheet.append(header)

    # Itera su ogni file di modello
    for model_file in model_files:
        model_path = os.path.join(MODELS_DIR, model_file)
        print(f"\n===== Inizio Test per il Modello: {model_file} =====")
        
        # Lista per contenere i risultati del modello corrente
        current_results_row = [model_file]

        # Itera sui 4 scenari
        for scenario_type in range(1, 5):
            env = Jaywalker()
            agent = QAgent(
                network=Lex_Q_Network,
                env=env,
                env_type=scenario_type,
                hidden=128
            )
            
            success_rate = agent.test_model(
                model_path=model_path, 
                num_episodes=NUM_TEST_EPISODES,
                render=False
            )
            # Aggiungi il risultato numerico (senza il simbolo %)
            current_results_row.append(success_rate)
        
        # Aggiungi la riga completa dei risultati al foglio Excel
        sheet.append(current_results_row)

    # Auto-formatta la larghezza delle colonne per una migliore leggibilità
    for col in sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Salva il file Excel
    try:
        workbook.save(RESULTS_EXCEL)
        print(f"\n===== Testing completato! =====")
        print(f"I risultati sono stati salvati nel file Excel: {RESULTS_EXCEL}")
    except IOError:
        print(f"Errore: Impossibile scrivere nel file {RESULTS_EXCEL}. Assicurati che non sia aperto.")
